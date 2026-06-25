import os, re, uuid, threading, zipfile, io, time, requests, logging, json, sys, hashlib
from datetime import datetime
from flask import Flask, request, jsonify, send_file, render_template
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pdfplumber

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# SELF-UPDATE
# ---------------------------------------------------------------------------
UPDATE_URL = "https://raw.githubusercontent.com/Bsgoncalves822/fam-app/main/app.py"
HASH_FILE  = os.path.join(BASE_DIR, 'data', 'update_hashes.json')
BRANCH     = "main"

def _md5(path):
    h = hashlib.md5()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''): h.update(chunk)
    return h.hexdigest()

def _load_hashes():
    try:
        with open(HASH_FILE) as f: return json.load(f)
    except: return {}

def _save_hashes(d):
    os.makedirs(os.path.dirname(HASH_FILE), exist_ok=True)
    with open(HASH_FILE, 'w') as f: json.dump(d, f)

def try_self_update():
    try:
        import subprocess
        result = subprocess.run(['git', 'fetch', '--quiet', 'origin', BRANCH],
                                cwd=BASE_DIR, capture_output=True, timeout=15)
        if result.returncode == 0:
            diff = subprocess.run(['git', 'diff', '--name-only', f'HEAD..origin/{BRANCH}'],
                                  cwd=BASE_DIR, capture_output=True, text=True, timeout=10,
                                  encoding='utf-8', errors='replace')
            changed = [l.strip() for l in diff.stdout.splitlines() if l.strip()]
            if changed:
                print(f"[UPDATE] {len(changed)} file(s) updated: {', '.join(changed)}")
                subprocess.run(['git', 'reset', '--hard', f'origin/{BRANCH}'],
                               cwd=BASE_DIR, capture_output=True, timeout=30)
                import shutil
                pycache = os.path.join(BASE_DIR, '__pycache__')
                if os.path.isdir(pycache): shutil.rmtree(pycache, ignore_errors=True)
                print("[UPDATE] Restarting...")
                os.execv(sys.executable, [sys.executable] + sys.argv)
            else:
                print("[UPDATE] Already up to date.")
            return
    except Exception as e:
        print(f"[UPDATE] Git unavailable ({e}), trying URL fallback...")
    try:
        hashes = _load_hashes()
        resp = requests.get(UPDATE_URL + f"?cb={int(time.time())}", timeout=10)
        if resp.status_code != 200: return
        remote_content = resp.content
        remote_hash = hashlib.md5(remote_content).hexdigest()
        local_hash  = _md5(__file__) if os.path.exists(__file__) else ''
        if remote_hash != local_hash and remote_hash != hashes.get('app.py', ''):
            with open(__file__, 'wb') as f: f.write(remote_content)
            hashes['app.py'] = remote_hash
            _save_hashes(hashes)
            import shutil
            pycache = os.path.join(BASE_DIR, '__pycache__')
            if os.path.isdir(pycache): shutil.rmtree(pycache, ignore_errors=True)
            print("[UPDATE] Restarting...")
            os.execv(sys.executable, [sys.executable] + sys.argv)
    except Exception as e:
        print(f"[UPDATE] Update failed ({e}), continuing with current version.")

# ---------------------------------------------------------------------------
# LOGGING
# ---------------------------------------------------------------------------
LOG_DIR = os.path.join(BASE_DIR, 'logs')
os.makedirs(LOG_DIR, exist_ok=True)
log_file = os.path.join(LOG_DIR, datetime.now().strftime('%Y-%m-%d') + '.log')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[logging.FileHandler(log_file, encoding='utf-8'), logging.StreamHandler()]
)
logger = logging.getLogger('fam')

# ---------------------------------------------------------------------------
# FLASK
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'uploads')
app.config['OUTPUT_FOLDER'] = os.path.join(BASE_DIR, 'outputs')
for d in [app.config['UPLOAD_FOLDER'], app.config['OUTPUT_FOLDER'],
          os.path.join(BASE_DIR, 'data')]:
    os.makedirs(d, exist_ok=True)

tasks = {}

# ---------------------------------------------------------------------------
# PDF EXTRACTION
# ---------------------------------------------------------------------------
FAM_CNPJ = '04957294000103'

def clean_cnpj(raw):
    return re.sub(r'[^\d]', '', str(raw))

def extract_pix(text):
    rec = {}
    m = re.search(r'Nome do destinat[aÃ¡]rio:\s*(.+)', text, re.IGNORECASE)
    if m: rec['destinatario_nome'] = m.group(1).strip()
    m = re.search(r'CNPJ do destinat[aÃ¡]rio:\s*([\d.\/\-]+)', text, re.IGNORECASE)
    if m:
        cnpj = clean_cnpj(m.group(1))
        if cnpj != FAM_CNPJ: rec['destinatario_cnpj'] = cnpj
    m = re.search(r'CPF do destinat[aÃ¡]rio:\s*([\*\d.\-]+)', text, re.IGNORECASE)
    if m: rec['destinatario_cpf'] = m.group(1).strip()
    m = re.search(r'Institui[cÃ§][aÃ£]o do destinat[aÃ¡]rio:\s*(.+)', text, re.IGNORECASE)
    if m: rec['banco'] = m.group(1).strip()
    m = re.search(r'Valor:\s*R\$\s*([\d.,]+)', text, re.IGNORECASE)
    if m: rec['valor'] = m.group(1).strip()
    m = re.search(r'Realizado em:\s*([\d/]+)', text, re.IGNORECASE)
    if m: rec['data'] = m.group(1).strip()
    m = re.search(r'Solicitante:\s*(.+)', text, re.IGNORECASE)
    if m: rec['solicitante'] = m.group(1).strip()
    m = re.search(r'ID da transa[cÃ§][aÃ£]o:\s*(\S+)', text, re.IGNORECASE)
    if m: rec['id_transacao'] = m.group(1).strip()
    m = re.search(r'N[uÃº]mero de Controle:\s*(\d+)', text, re.IGNORECASE)
    if m: rec['num_controle'] = m.group(1).strip()
    # label like RESCISÃƒO sitting between header and Valor line
    m = re.search(r'Comprovante de Pagamento Pix\s*\n([A-ZÃÃ‰ÃÃ“ÃšÃƒÃ•Ã‡ ]{3,})\n', text, re.IGNORECASE)
    if m:
        label = m.group(1).strip()
        if label.upper() not in ('VALOR', 'REALIZADO'):
            rec['label'] = label
    rec['tipo'] = 'PIX'
    return rec

def extract_boleto(text):
    rec = {}
    for pat in [r'Raz[aÃ£]o Social do Benefici[aÃ¡]rio:\s*(.+)',
                r'Nome Fantasia do Benefici[aÃ¡]rio:\s*(.+)']:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            rec['destinatario_nome'] = m.group(1).strip()
            break
    m = re.search(r'CPF/CNPJ do Benefici[aÃ¡]rio:\s*([\d.\/\-]+)', text, re.IGNORECASE)
    if m:
        cnpj = clean_cnpj(m.group(1))
        if cnpj != FAM_CNPJ: rec['destinatario_cnpj'] = cnpj
    m = re.search(r'Institui[cÃ§][aÃ£]o Emissora:\s*(.+)', text, re.IGNORECASE)
    if m: rec['banco'] = m.group(1).strip()
    m = re.search(r'Valor do T[iÃ­]tulo \(R\$\):\s*([\d.,]+)', text, re.IGNORECASE)
    if not m: m = re.search(r'Valor\s*\(R\$\):\s*([\d.,]+)', text, re.IGNORECASE)
    if m: rec['valor'] = m.group(1).strip()
    m = re.search(r'Data do Pagamento:\s*([\d/]+)', text, re.IGNORECASE)
    if not m: m = re.search(r'Data da Transa[cÃ§][aÃ£]o:\s*([\d/]+)', text, re.IGNORECASE)
    if m: rec['data'] = m.group(1).strip()
    m = re.search(r'Solicitante:\s*(.+)', text, re.IGNORECASE)
    if m: rec['solicitante'] = m.group(1).strip()
    m = re.search(r'N[uÃº]mero de Controle:\s*(\d+)', text, re.IGNORECASE)
    if m: rec['num_controle'] = m.group(1).strip()
    m = re.search(r'Data de Vencimento:\s*([\d/]+)', text, re.IGNORECASE)
    if m: rec['data_vencimento'] = m.group(1).strip()
    rec['tipo'] = 'BOLETO'
    return rec

def extract_pdf_data(pdf_bytes):
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = '\n'.join(p.extract_text() or '' for p in pdf.pages[:3])
    except:
        return {'tipo': 'ERRO', 'obs': 'PDF ilegÃ­vel'}
    if not text.strip():
        return {'tipo': 'IMAGEM', 'obs': 'PDF sem texto extraÃ­vel'}
    if re.search(r'Comprovante de Pagamento Pix', text, re.IGNORECASE):
        return extract_pix(text)
    elif re.search(r'Pagar Boletos|Benefici[aÃ¡]rio', text, re.IGNORECASE):
        return extract_boleto(text)
    else:
        rec = {'tipo': 'OUTRO'}
        m = re.search(r'Valor[:\s]+R\$\s*([\d.,]+)', text, re.IGNORECASE)
        if m: rec['valor'] = m.group(1).strip()
        return rec

# ---------------------------------------------------------------------------
# FILENAME PARSER
# ---------------------------------------------------------------------------
def fix_year(y):
    y = y[:4]
    if y.startswith('200') and int(y) > 2030: y = '20' + y[2:]
    return y

def parse_amount_str(raw):
    if raw is None: return None
    raw = str(raw).strip().lstrip('-').strip()
    raw = re.sub(r'[^\d,.]', '', raw)
    if ',' in raw:
        parts = raw.rsplit(',', 1)
        integer_part = parts[0].replace('.', '').replace(',', '')
        try: return float(f"{integer_part}.{parts[1]}")
        except: return None
    return None

def parse_filename(filename):
    name = re.sub(r'\.pdf$', '', filename, flags=re.IGNORECASE).strip()
    name = re.sub(r'^[a-zA-Z]+(?=\d{1,2}[-.\s]\d{2}[-.\s])', '', name)
    name = re.sub(r'\s*\(VENC[^)]*\)\s*$', '', name, flags=re.IGNORECASE)
    date_m = re.match(r'^(\d{1,2})[\s.\-]+(\d{2})[\s.\-+]+(\d{4,5})[\s.\-]*\s*', name)
    if not date_m:
        return None, None, None
    date_str  = f"{date_m.group(1).zfill(2)}/{date_m.group(2)}/{fix_year(date_m.group(3))}"
    remainder = name[date_m.end():].strip().lstrip('-â€“').strip()
    amount_m  = re.search(r'[-â€“]\s*([\d.,]+,\d{2,})\s*$', remainder)
    if amount_m:
        amount    = parse_amount_str(amount_m.group(1))
        desc      = remainder[:amount_m.start()].strip().rstrip('-â€“').strip()
    else:
        amount = None
        desc   = remainder.strip()
    return date_str, desc, amount

# ---------------------------------------------------------------------------
# EXCEL BUILDER  â€” single clean sheet, easy to paste
# ---------------------------------------------------------------------------
HEADERS = [
    "Data",
    "Valor (R$)",
    "Tipo",
    "Label",
    "DestinatÃ¡rio",
    "CNPJ DestinatÃ¡rio",
    "CPF DestinatÃ¡rio",
    "Banco DestinatÃ¡rio",
    "Solicitante",
    "NÂº Controle",
    "ID TransaÃ§Ã£o",
    "Data Vencimento",
    "DescriÃ§Ã£o (filename)",
    "Arquivo",
    "Obs",
]
WIDTHS = [13, 14, 10, 14, 40, 20, 18, 32, 18, 16, 38, 16, 45, 58, 25]

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name="Arial", bold=True, size=9, color="FFFFFF")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
HDR_BORDER = Border(bottom=Side(style="medium", color="2E75B6"))

ODD_FILL  = PatternFill("solid", fgColor="FFFFFF")
EVEN_FILL = PatternFill("solid", fgColor="F2F7FC")
ERR_FILL  = PatternFill("solid", fgColor="FFF2CC")
ROW_BORDER = Border(bottom=Side(style="thin", color="D9E1F2"))

def build_excel(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "COMPROVANTES"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for ci, (hdr, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border    = HDR_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 20

    for ri, row in enumerate(rows, 2):
        fill = ERR_FILL if row.get('obs') else (EVEN_FILL if ri % 2 == 0 else ODD_FILL)

        values = [
            row.get('data', ''),
            row.get('valor'),
            row.get('tipo', ''),
            row.get('label', ''),
            row.get('destinatario_nome', ''),
            row.get('destinatario_cnpj', ''),
            row.get('destinatario_cpf', ''),
            row.get('banco', ''),
            row.get('solicitante', ''),
            row.get('num_controle', ''),
            row.get('id_transacao', ''),
            row.get('data_vencimento', ''),
            row.get('desc_filename', ''),
            row.get('arquivo', ''),
            row.get('obs', ''),
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center")
            cell.border    = ROW_BORDER
            cell.font      = Font(name="Arial", size=9)
            if ci == 2:  # Valor
                cell.number_format = '#,##0.00'
        ws.row_dimensions[ri].height = 15

    # auto-filter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    wb.save(out_path)

# ---------------------------------------------------------------------------
# TASK RUNNER
# ---------------------------------------------------------------------------
def run_parse(task_id, zip_path):
    def log(msg):
        tasks[task_id]['log'].append(msg)
        logger.info(msg)

    try:
        log("Lendo ZIP...")
        with zipfile.ZipFile(zip_path, 'r') as zf:
            all_names = [n for n in zf.namelist() if n.lower().endswith('.pdf')]
        log(f"{len(all_names)} comprovantes encontrados. Extraindo PDFs...")

        rows = []
        with zipfile.ZipFile(zip_path, 'r') as zf:
            for i, fname in enumerate(all_names):
                if (i + 1) % 200 == 0:
                    log(f"  {i+1}/{len(all_names)}...")
                basename = os.path.basename(fname)
                date_fn, desc_fn, amount_fn = parse_filename(basename)
                try:
                    pdf_data = extract_pdf_data(zf.read(fname))
                except Exception as e:
                    pdf_data = {'tipo': 'ERRO', 'obs': str(e)}

                # prefer PDF-extracted data, fall back to filename
                valor_str = pdf_data.get('valor')
                if valor_str:
                    try:
                        valor = float(valor_str.replace('.', '').replace(',', '.'))
                    except:
                        valor = amount_fn
                else:
                    valor = amount_fn

                rows.append({
                    'arquivo':           basename,
                    'tipo':              pdf_data.get('tipo', ''),
                    'label':             pdf_data.get('label', ''),
                    'data':              pdf_data.get('data') or date_fn or '',
                    'valor':             valor,
                    'destinatario_nome': pdf_data.get('destinatario_nome', ''),
                    'destinatario_cnpj': pdf_data.get('destinatario_cnpj', ''),
                    'destinatario_cpf':  pdf_data.get('destinatario_cpf', ''),
                    'banco':             pdf_data.get('banco', ''),
                    'solicitante':       pdf_data.get('solicitante', ''),
                    'num_controle':      pdf_data.get('num_controle', ''),
                    'id_transacao':      pdf_data.get('id_transacao', ''),
                    'data_vencimento':   pdf_data.get('data_vencimento', ''),
                    'desc_filename':     desc_fn or '',
                    'obs':               pdf_data.get('obs', ''),
                })

        ok     = sum(1 for r in rows if not r['obs'])
        flagged = len(rows) - ok
        log(f"ExtraÃ§Ã£o: {ok} OK Â· {flagged} com observaÃ§Ã£o")
        log("Gerando Excel...")

        out_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{task_id}.xlsx")
        build_excel(rows, out_path)

        tasks[task_id]['status'] = 'DONE'
        tasks[task_id]['file']   = out_path
        log(f"Pronto! {len(rows)} linhas â†’ FAM_Comprovantes.xlsx")

    except Exception as e:
        import traceback
        tasks[task_id]['status'] = 'ERROR'
        tasks[task_id]['log'].append(f"ERRO: {str(e)}")
        tasks[task_id]['log'].append(traceback.format_exc())
        logger.error(f"Run failed: {e}")

# ---------------------------------------------------------------------------
# ROUTES
# ---------------------------------------------------------------------------
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload/zip', methods=['POST'])
def upload_zip():
    if 'file' not in request.files: return jsonify({"error": "No file"}), 400
    f = request.files['file']
    path = os.path.join(app.config['UPLOAD_FOLDER'], f"comp_{uuid.uuid4().hex}.zip")
    f.save(path)
    return jsonify({"path": path})

@app.route('/processar', methods=['POST'])
def processar():
    data     = request.json
    zip_path = data.get('zip_path')
    if not zip_path: return jsonify({"error": "Missing zip"}), 400
    task_id = uuid.uuid4().hex
    tasks[task_id] = {"status": "PENDING", "log": [], "file": None}
    threading.Thread(target=run_parse, args=(task_id, zip_path)).start()
    return jsonify({"task_id": task_id})

@app.route('/status/<task_id>')
def status(task_id):
    if task_id not in tasks: return jsonify({"error": "Invalid task"}), 404
    return jsonify(tasks[task_id])

@app.route('/download/<task_id>')
def download(task_id):
    if task_id not in tasks or tasks[task_id]['status'] != 'DONE':
        return jsonify({"error": "Not ready"}), 400
    return send_file(tasks[task_id]['file'], as_attachment=True,
                     download_name='FAM_Comprovantes.xlsx')

# ---------------------------------------------------------------------------
# ENTRY POINT
# ---------------------------------------------------------------------------
if __name__ == '__main__':
    import threading
    threading.Thread(target=try_self_update, daemon=True).start()
    logger.info("FAM App (Comprovantes Parser) starting on port 5002")
    app.run(host='0.0.0.0', port=5002, debug=False)

